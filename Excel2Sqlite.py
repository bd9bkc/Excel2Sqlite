from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import chardet
import sqlite3

#导入Excel文件
dest_filename = '选题名单.xlsx'
excel = load_workbook(dest_filename)
excel_wb = excel.active

#获取最大行数
print ("Max_row",excel_wb.max_row)
MaxiumRow = excel_wb.max_row+1
count_mianshi =0;
count_kaiti = 0;
count_cansai = 0;

#连接到数据库
dbconn = sqlite3.connect("stu.db")

#创建一个游标
cursor = dbconn.cursor()

#创建stuinfo 表
#sql = ""
#cursor.execute(sql)

select = "select * from stuinfo limit 1 offset (select count(*) - 1  from stuinfo)"
         #"select * from stuinfo order by 'index' desc limit 0,1"
cursor.execute(select)
values = cursor.fetchone()
#print(values)
MaxiumIndex = str(int(values[0]) + 1)
#print (MaxiumIndex)


#按行遍历excel文件
for row in range(2,MaxiumRow):
    #print(excel_wb.cell(row=row,column=1).value)
    #读取ID
    stu_no=excel_wb.cell(row=row,column=1).value
    #print("ID",stu_no)
    #读取Name
    stu_name=excel_wb.cell(row=row,column=2).value
    #读取class
    stu_class=excel_wb.cell(row=row,column=5).value

    #读取 Status 和 Title， 注意可能为空
    detial_info=excel_wb.cell(row=row,column=9).value

    if not(detial_info is None) :
        #detial_info.replace("，",",")
        #detial_info.replace(".",",")
        #detial_info.replace(" ","")

        #分割信息，得到 status 和 title
        if detial_info.count(",") == 1:
            stu_type = detial_info.split(",", 2)[0]
            title= detial_info.split(",",2)[1]
        elif detial_info.count(",") == 2:
            stu_type = detial_info.split(",", 3)[0]
            title = detial_info.split(",", 3)[1] + "：" + detial_info.split(",", 3)[2]
        else:
            print ("ID:",stu_no,"信息错误")
            stu_type = "Error"
            title="Error"
        #print(stu_type,title)

        selectbyid = "select * from stuinfo where id = ?"
        # select = "select * from stuinfo"
        # cursor.execute(select)
        cursor.execute(selectbyid,(stu_no,))
        value = cursor.fetchone()
        #print(value)

        if (value is None):
            #如果没有，则新增加一条记录
            sqlinsert = "INSERT INTO stuinfo('index',id,name,class,status,title) VALUES (?,?,?,?,?,?)"
            data = (MaxiumIndex,stu_no,stu_name,stu_class,stu_type,title)
            print("Insert ID",stu_no,"to Index:",MaxiumIndex)
            MaxiumIndex = str(int(MaxiumIndex) + 1)

            #print(sqlinsert)
            #print(MaxiumIndex)
            #print(data)
            cursor.execute(sqlinsert, data)
        else:
            #如果存在，则修改原来记录
            #print(value)
            sqlupdate = "update stuinfo set status=?, title=? where id =?"
            cursor.execute(sqlupdate, (stu_type,title,stu_no))
            print("Update ID",stu_no,"at Index:",value[0])
            #print(sqlupdate)
        dbconn.commit()

dbconn.commit()
dbconn.close()